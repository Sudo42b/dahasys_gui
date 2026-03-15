"""
MINIAS - Probe Testing System
Visual Basic 6.0에서 Python으로 재구현
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Tuple
import statistics
import threading
import queue
import os
import configparser

# 시리얼 통신
try:
    import serial
    import serial.tools.list_ports

    SERIAL_AVAILABLE = True
except ImportError:
    SERIAL_AVAILABLE = False
    print("Warning: pyserial not installed. Serial communication disabled.")

# Excel 출력
try:
    from openpyxl import Workbook, load_workbook

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export disabled.")

# PDF 인증서 출력
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.units import mm

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("Warning: reportlab not installed. PDF certificate disabled.")


# =============================================================================
# 데이터 모델
# =============================================================================


@dataclass
class TestResult:
    """테스트 결과 데이터"""

    id_col: int = 0
    date: datetime = field(default_factory=datetime.now)
    code: str = ""
    serial_number: str = ""
    operator: str = ""
    test_type: str = "ST"
    result: str = ""
    mean_sigma: float = 0.0
    mean_range: float = 0.0
    worst_sigma: float = 0.0
    worst_range: float = 0.0
    mean_sigma_limit: float = 0.0
    mean_range_limit: float = 0.0
    worst_range_limit: float = 0.0
    second_test: str = "N"


@dataclass
class AxisResult:
    """축별 테스트 결과"""

    id_col: int = 0
    axis: int = 0
    direction: str = ""
    sigma: float = 0.0
    range_val: float = 0.0
    result: str = ""
    ncycles: int = 0
    second_test: str = "N"


@dataclass
class CodeInfo:
    """프로브 코드 정보"""

    code: str = ""
    naxis: int = 4
    probe_type: str = ""
    x_plus_dir: int = 1
    x_minus_dir: int = 1
    y_plus_dir: int = 1
    y_minus_dir: int = 1
    z_minus_dir: int = 1


@dataclass
class SetupInfo:
    """테스트 설정 정보"""

    test_type: str = "ST"
    ncycles: int = 100
    naxis: int = 4
    nworstdatums: int = 1
    sigma_test: bool = True
    range_test: bool = True


@dataclass
class LimitInfo:
    """한계값 정보"""

    test_type: str = "ST"
    mean_sigma: float = 0.0
    mean_range: float = 0.0
    worst_range: float = 0.0
    mean_range_performance: float = 0.0
    worst_range_performance: float = 0.0
    mean_range_second: float = 0.0
    worst_range_second: float = 0.0


# =============================================================================
# 데이터베이스 모듈
# =============================================================================


class MiniasDatabase:
    """SQLite 데이터베이스 관리"""

    def __init__(self, db_path: str = "minias.db"):
        self.db_path = db_path
        self.conn: Optional[sqlite3.Connection] = None

    def connect(self):
        """데이터베이스 연결"""
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._init_tables()

    def close(self):
        """연결 종료"""
        if self.conn:
            self.conn.close()
            self.conn = None

    def _init_tables(self):
        """테이블 초기화"""
        cursor = self.conn.cursor()

        # OPERATORS 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS OPERATORS (
                OPERATOR VARCHAR(20) PRIMARY KEY
            )
        """)

        # CODES 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS CODES (
                CODE VARCHAR(11) PRIMARY KEY,
                NAXIS INT DEFAULT 4,
                PROBE_TYPE VARCHAR(30),
                XPLUS_DIR SMALLINT DEFAULT 1,
                XMINUS_DIR SMALLINT DEFAULT 1,
                YPLUS_DIR SMALLINT DEFAULT 1,
                YMINUS_DIR SMALLINT DEFAULT 1,
                ZMINUS_DIR SMALLINT DEFAULT 1
            )
        """)

        # SETUP 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS SETUP (
                TEST_TYPE VARCHAR(2) PRIMARY KEY,
                NCYCLES INT DEFAULT 100,
                NAXIS SMALLINT DEFAULT 4,
                NWORSTDATUMS SMALLINT DEFAULT 1,
                SIGMA_TEST BOOLEAN DEFAULT 1,
                RANGE_TEST BOOLEAN DEFAULT 1
            )
        """)

        # LIMITS 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS LIMITS (
                TEST_TYPE VARCHAR(2) PRIMARY KEY,
                MEAN_SIGMA REAL,
                MEAN_RANGE REAL,
                WORST_RANGE REAL,
                MEAN_RANGE_PERFORMANCE REAL,
                WORST_RANGE_PERFORMANCE REAL,
                MEAN_RANGE_SECOND REAL,
                WORST_RANGE_SECOND REAL
            )
        """)

        # TEST_RESULTS 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS TEST_RESULTS (
                ID_COL INTEGER PRIMARY KEY AUTOINCREMENT,
                DATE DATETIME,
                CODE VARCHAR(20),
                SERIAL_NUMBER VARCHAR(10),
                OPERATOR VARCHAR(20),
                TEST VARCHAR(2),
                RESULT VARCHAR(2),
                MEAN_SIGMA REAL,
                MEAN_RANGE REAL,
                WORST_SIGMA REAL,
                WORST_RANGE REAL,
                MEAN_SIGMA_LIMIT REAL,
                MEAN_RANGE_LIMIT REAL,
                WORST_RANGE_LIMIT REAL,
                SECOND_TEST VARCHAR(1) DEFAULT 'N'
            )
        """)

        # TEST_AXIS_RESULTS 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS TEST_AXIS_RESULTS (
                ID_COL INT,
                AXIS SMALLINT,
                DIR VARCHAR(2),
                SIGMA REAL,
                RANGE REAL,
                RESULT VARCHAR(2),
                NCYCLES INT,
                SECOND_TEST VARCHAR(1) DEFAULT 'N',
                PRIMARY KEY (ID_COL, AXIS)
            )
        """)

        # TEST_SAMPLES 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS TEST_SAMPLES (
                ID_COL INT,
                AXIS SMALLINT,
                CYCLE SMALLINT,
                VALUE REAL,
                SECOND_TEST VARCHAR(1) DEFAULT 'N',
                WORST_DATUM VARCHAR(1) DEFAULT 'N',
                PRIMARY KEY (ID_COL, AXIS, CYCLE)
            )
        """)

        # MEASURES 테이블 (현재 측정값)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS MEASURES (
                AXIS SMALLINT,
                CYCLE INT,
                VALUE REAL,
                SECOND_TEST VARCHAR(1) DEFAULT 'N',
                WORST_DATUM VARCHAR(1) DEFAULT 'N',
                PRIMARY KEY (AXIS, CYCLE)
            )
        """)

        # EXCEL_SETUP 테이블
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS EXCEL_SETUP (
                ID INTEGER PRIMARY KEY,
                PROBE_TYPE VARCHAR(10),
                CODE VARCHAR(10),
                SERIAL_NUMBER VARCHAR(10),
                OPERATOR VARCHAR(10),
                MEAN_SIGMA VARCHAR(10),
                MEAN_RANGE VARCHAR(10),
                WORST_RANGE VARCHAR(10),
                LIMIT_MEAN_RANGE VARCHAR(10),
                LIMIT_WORST_RANGE VARCHAR(10),
                RESULT VARCHAR(10),
                ASSE1_SIGMA VARCHAR(10),
                ASSE1_RANGE VARCHAR(10),
                ASSE1_START VARCHAR(10),
                DELTA_SECOND SMALLINT DEFAULT 0
            )
        """)

        self.conn.commit()

        # 기본 데이터 삽입
        self._insert_default_data()

    def _insert_default_data(self):
        """기본 데이터 삽입"""
        cursor = self.conn.cursor()

        # 기본 Setup 추가
        cursor.execute("""
            INSERT OR IGNORE INTO SETUP (TEST_TYPE, NCYCLES, NAXIS, NWORSTDATUMS, SIGMA_TEST, RANGE_TEST)
            VALUES ('ST', 100, 4, 1, 1, 1)
        """)

        # 기본 Limits 추가
        cursor.execute("""
            INSERT OR IGNORE INTO LIMITS (TEST_TYPE, MEAN_SIGMA, MEAN_RANGE, WORST_RANGE)
            VALUES ('ST', 0.005, 0.010, 0.015)
        """)

        self.conn.commit()

    # --- OPERATORS ---
    def get_operators(self) -> List[str]:
        """작업자 목록 조회"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT OPERATOR FROM OPERATORS ORDER BY OPERATOR")
        return [row["OPERATOR"] for row in cursor.fetchall()]

    def add_operator(self, operator: str):
        """작업자 추가"""
        cursor = self.conn.cursor()
        cursor.execute(
            "INSERT OR IGNORE INTO OPERATORS (OPERATOR) VALUES (?)", (operator,)
        )
        self.conn.commit()

    # --- CODES ---
    def get_codes(self) -> List[CodeInfo]:
        """코드 목록 조회"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM CODES ORDER BY CODE")
        results = []

        # 컬럼명 가져오기
        col_names = (
            [desc[0] for desc in cursor.description] if cursor.description else []
        )

        for row in cursor.fetchall():

            def safe_get(name, alt_name=None, default=None):
                """컬럼이 존재하면 값 반환"""
                if name in col_names:
                    val = row[name]
                    return val if val is not None else default
                if alt_name and alt_name in col_names:
                    val = row[alt_name]
                    return val if val is not None else default
                return default

            results.append(
                CodeInfo(
                    code=safe_get("CODE", default=""),
                    naxis=int(safe_get("NAXIS", default=4) or 4),
                    probe_type=safe_get("PROBE_TYPE", default=""),
                    x_plus_dir=int(safe_get("X_PLUS_DIR", "XPLUS_DIR", 1) or 1),
                    x_minus_dir=int(safe_get("X_MINUS_DIR", "XMINUS_DIR", 1) or 1),
                    y_plus_dir=int(safe_get("Y_PLUS_DIR", "YPLUS_DIR", 1) or 1),
                    y_minus_dir=int(safe_get("Y_MINUS_DIR", "YMINUS_DIR", 1) or 1),
                    z_minus_dir=int(safe_get("Z_MINUS_DIR", "ZMINUS_DIR", 1) or 1),
                )
            )
        return results

    def get_code_list(self) -> List[str]:
        """코드 문자열 목록 조회"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT CODE FROM CODES ORDER BY CODE")
        return [row["CODE"] for row in cursor.fetchall()]

    def get_code_info(self, code: str) -> Optional[CodeInfo]:
        """특정 코드 정보 조회"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM CODES WHERE CODE = ?", (code,))
        row = cursor.fetchone()
        if row:
            col_names = (
                [desc[0] for desc in cursor.description] if cursor.description else []
            )

            def safe_get(name, alt_name=None, default=None):
                if name in col_names:
                    val = row[name]
                    return val if val is not None else default
                if alt_name and alt_name in col_names:
                    val = row[alt_name]
                    return val if val is not None else default
                return default

            return CodeInfo(
                code=safe_get("CODE", default=""),
                naxis=int(safe_get("NAXIS", default=4) or 4),
                probe_type=safe_get("PROBE_TYPE", default=""),
            )
        return None

    def add_code(self, code_info: CodeInfo):
        """코드 추가"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            INSERT OR REPLACE INTO CODES
            (CODE, NAXIS, PROBE_TYPE, XPLUS_DIR, XMINUS_DIR, YPLUS_DIR, YMINUS_DIR, ZMINUS_DIR)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                code_info.code,
                code_info.naxis,
                code_info.probe_type,
                code_info.x_plus_dir,
                code_info.x_minus_dir,
                code_info.y_plus_dir,
                code_info.y_minus_dir,
                code_info.z_minus_dir,
            ),
        )
        self.conn.commit()

    def delete_code(self, code: str):
        """코드 삭제"""
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM CODES WHERE CODE = ?", (code,))
        self.conn.commit()

    def delete_test_result(self, id_col: int):
        """테스트 결과 및 관련 데이터 삭제"""
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM TEST_SAMPLES WHERE ID_COL = ?", (id_col,))
        cursor.execute("DELETE FROM TEST_AXIS_RESULTS WHERE ID_COL = ?", (id_col,))
        cursor.execute("DELETE FROM TEST_RESULTS WHERE ID_COL = ?", (id_col,))
        self.conn.commit()

    def search_by_serial_number(self, serial_number: str) -> List[Tuple[int, str, str]]:
        """시리얼 번호로 테스트 결과 검색"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            SELECT ID_COL, CODE, SERIAL_NUMBER FROM TEST_RESULTS
            WHERE SERIAL_NUMBER LIKE ? ORDER BY ID_COL DESC
        """,
            (f"%{serial_number}%",),
        )
        return [
            (row["ID_COL"], row["CODE"] or "", row["SERIAL_NUMBER"] or "")
            for row in cursor.fetchall()
        ]

    # --- SETUP ---
    def get_setup(self, test_type: str = "ST") -> Optional[SetupInfo]:
        """설정 조회"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM SETUP WHERE TEST_TYPE = ?", (test_type,))
        row = cursor.fetchone()
        if row:
            return SetupInfo(
                test_type=row["TEST_TYPE"],
                ncycles=row["NCYCLES"],
                naxis=row["NAXIS"],
                nworstdatums=row["NWORSTDATUMS"],
                sigma_test=bool(row["SIGMA_TEST"]),
                range_test=bool(row["RANGE_TEST"]),
            )
        return None

    def save_setup(self, setup: SetupInfo):
        """설정 저장"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            INSERT OR REPLACE INTO SETUP
            (TEST_TYPE, NCYCLES, NAXIS, NWORSTDATUMS, SIGMA_TEST, RANGE_TEST)
            VALUES (?, ?, ?, ?, ?, ?)
        """,
            (
                setup.test_type,
                setup.ncycles,
                setup.naxis,
                setup.nworstdatums,
                setup.sigma_test,
                setup.range_test,
            ),
        )
        self.conn.commit()

    # --- LIMITS ---
    def get_limits(self, test_type: str = "ST") -> Optional[LimitInfo]:
        """한계값 조회"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM LIMITS WHERE TEST_TYPE = ?", (test_type,))
        row = cursor.fetchone()
        if row:
            # 컬럼 이름 목록 가져오기 (동적으로 처리)
            col_names = [desc[0] for desc in cursor.description]

            def safe_get(name, default=0.0):
                """컬럼이 존재하면 값 반환, 없으면 기본값"""
                if name in col_names:
                    val = row[name]
                    return val if val is not None else default
                return default

            return LimitInfo(
                test_type=safe_get("TEST_TYPE", "ST"),
                mean_sigma=safe_get("MEAN_SIGMA"),
                mean_range=safe_get("MEAN_RANGE"),
                worst_range=safe_get("WORST_RANGE"),
                mean_range_performance=safe_get("MEAN_RANGE_PERFORMANCE"),
                worst_range_performance=safe_get("WORST_RANGE_PERFORMANCE"),
                mean_range_second=safe_get("MEAN_RANGE_SECOND"),
                worst_range_second=safe_get("WORST_RANGE_SECOND"),
            )
        return None

    def save_limits(self, limits: LimitInfo):
        """한계값 저장"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            INSERT OR REPLACE INTO LIMITS
            (TEST_TYPE, MEAN_SIGMA, MEAN_RANGE, WORST_RANGE,
             MEAN_RANGE_PERFORMANCE, WORST_RANGE_PERFORMANCE,
             MEAN_RANGE_SECOND, WORST_RANGE_SECOND)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                limits.test_type,
                limits.mean_sigma,
                limits.mean_range,
                limits.worst_range,
                limits.mean_range_performance,
                limits.worst_range_performance,
                limits.mean_range_second,
                limits.worst_range_second,
            ),
        )
        self.conn.commit()

    # --- TEST_RESULTS ---
    def save_test_result(self, result: TestResult) -> int:
        """테스트 결과 저장"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            INSERT INTO TEST_RESULTS
            (DATE, CODE, SERIAL_NUMBER, OPERATOR, TEST, RESULT,
             MEAN_SIGMA, MEAN_RANGE, WORST_SIGMA, WORST_RANGE,
             MEAN_SIGMA_LIMIT, MEAN_RANGE_LIMIT, WORST_RANGE_LIMIT, SECOND_TEST)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                result.date.isoformat(),
                result.code,
                result.serial_number,
                result.operator,
                result.test_type,
                result.result,
                result.mean_sigma,
                result.mean_range,
                result.worst_sigma,
                result.worst_range,
                result.mean_sigma_limit,
                result.mean_range_limit,
                result.worst_range_limit,
                result.second_test,
            ),
        )
        self.conn.commit()
        return cursor.lastrowid

    def get_test_result(self, id_col: int) -> Optional[TestResult]:
        """특정 테스트 결과 조회"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            SELECT * FROM TEST_RESULTS WHERE ID_COL = ?
        """,
            (id_col,),
        )
        row = cursor.fetchone()
        if row:
            col_names = (
                [desc[0] for desc in cursor.description] if cursor.description else []
            )

            def safe_get(name, default=None):
                if name in col_names:
                    val = row[name]
                    return val if val is not None else default
                return default

            date_str = safe_get("DATE")
            try:
                date_val = (
                    datetime.fromisoformat(date_str) if date_str else datetime.now()
                )
            except (ValueError, TypeError):
                date_val = datetime.now()

            return TestResult(
                id_col=safe_get("ID_COL", 0),
                date=date_val,
                code=safe_get("CODE", ""),
                serial_number=safe_get("SERIAL_NUMBER", ""),
                operator=safe_get("OPERATOR", ""),
                test_type=safe_get("TEST", "ST"),
                result=safe_get("RESULT", ""),
                mean_sigma=float(safe_get("MEAN_SIGMA", 0.0) or 0.0),
                mean_range=float(safe_get("MEAN_RANGE", 0.0) or 0.0),
                worst_sigma=float(safe_get("WORST_SIGMA", 0.0) or 0.0),
                worst_range=float(safe_get("WORST_RANGE", 0.0) or 0.0),
                mean_sigma_limit=float(safe_get("MEAN_SIGMA_LIMIT", 0.0) or 0.0),
                mean_range_limit=float(safe_get("MEAN_RANGE_LIMIT", 0.0) or 0.0),
                worst_range_limit=float(safe_get("WORST_RANGE_LIMIT", 0.0) or 0.0),
                second_test=safe_get("SECOND_TEST", "N"),
            )
        return None

    def get_last_id(self) -> int:
        """마지막 ID 조회"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT MAX(ID_COL) as max_id FROM TEST_RESULTS")
        row = cursor.fetchone()
        return row["max_id"] if row and row["max_id"] else 0

    def get_recent_ids(self, limit: int = 50) -> List[Tuple[int, str, str]]:
        """최근 테스트 ID 목록 조회 (ID, Code, Serial)"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            SELECT ID_COL, CODE, SERIAL_NUMBER FROM TEST_RESULTS
            ORDER BY ID_COL DESC LIMIT ?
        """,
            (limit,),
        )
        return [
            (row["ID_COL"], row["CODE"] or "", row["SERIAL_NUMBER"] or "")
            for row in cursor.fetchall()
        ]

    # --- TEST_AXIS_RESULTS ---
    def save_axis_result(self, result: AxisResult):
        """축별 결과 저장"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            INSERT OR REPLACE INTO TEST_AXIS_RESULTS
            (ID_COL, AXIS, DIR, SIGMA, RANGE, RESULT, NCYCLES, SECOND_TEST)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                result.id_col,
                result.axis,
                result.direction,
                result.sigma,
                result.range_val,
                result.result,
                result.ncycles,
                result.second_test,
            ),
        )
        self.conn.commit()

    def get_axis_results(self, id_col: int) -> List[AxisResult]:
        """축별 결과 조회"""
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT * FROM TEST_AXIS_RESULTS WHERE ID_COL = ? ORDER BY AXIS", (id_col,)
        )
        results = []
        for row in cursor.fetchall():
            # axis 값을 int로 변환 (문자열일 수 있음)
            axis_val = row["AXIS"]
            try:
                axis_val = int(axis_val) if axis_val is not None else 0
            except (ValueError, TypeError):
                axis_val = 0

            results.append(
                AxisResult(
                    id_col=row["ID_COL"],
                    axis=axis_val,
                    direction=row["DIR"] or "",
                    sigma=float(row["SIGMA"] or 0.0),
                    range_val=float(row["RANGE"] or 0.0),
                    result=row["RESULT"] or "",
                    ncycles=int(row["NCYCLES"] or 0),
                    second_test=row["SECOND_TEST"] or "N",
                )
            )
        return results

    # --- TEST_SAMPLES ---
    def save_sample(
        self,
        id_col: int,
        axis: int,
        cycle: int,
        value: float,
        second_test: str = "N",
        worst_datum: str = "N",
    ):
        """샘플 데이터 저장"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            INSERT OR REPLACE INTO TEST_SAMPLES
            (ID_COL, AXIS, CYCLE, VALUE, SECOND_TEST, WORST_DATUM)
            VALUES (?, ?, ?, ?, ?, ?)
        """,
            (id_col, axis, cycle, value, second_test, worst_datum),
        )
        self.conn.commit()

    def get_samples(self, id_col: int, axis: int = None) -> List[Tuple]:
        """샘플 데이터 조회"""
        cursor = self.conn.cursor()
        if axis:
            cursor.execute(
                """
                SELECT * FROM TEST_SAMPLES
                WHERE ID_COL = ? AND AXIS = ? ORDER BY CYCLE
            """,
                (id_col, axis),
            )
        else:
            cursor.execute(
                """
                SELECT * FROM TEST_SAMPLES WHERE ID_COL = ? ORDER BY AXIS, CYCLE
            """,
                (id_col,),
            )
        return cursor.fetchall()

    # --- MEASURES (현재 측정값) ---
    def clear_measures(self):
        """현재 측정값 초기화"""
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM MEASURES")
        self.conn.commit()

    def save_measure(
        self,
        axis: int,
        cycle: int,
        value: float,
        second_test: str = "N",
        worst_datum: str = "N",
    ):
        """현재 측정값 저장"""
        cursor = self.conn.cursor()
        cursor.execute(
            """
            INSERT OR REPLACE INTO MEASURES (AXIS, CYCLE, VALUE, SECOND_TEST, WORST_DATUM)
            VALUES (?, ?, ?, ?, ?)
        """,
            (axis, cycle, value, second_test, worst_datum),
        )
        self.conn.commit()

    def get_measures(self, axis: int = None) -> List[Tuple]:
        """현재 측정값 조회"""
        cursor = self.conn.cursor()
        if axis:
            cursor.execute(
                "SELECT * FROM MEASURES WHERE AXIS = ? ORDER BY CYCLE", (axis,)
            )
        else:
            cursor.execute("SELECT * FROM MEASURES ORDER BY AXIS, CYCLE")
        return cursor.fetchall()


# =============================================================================
# 시리얼 통신 모듈
# =============================================================================


class SerialCommunicator:
    """시리얼 통신 관리"""

    def __init__(self, port: str = "COM1", baudrate: int = 9600):
        self.port = port
        self.baudrate = baudrate
        self.serial: Optional[serial.Serial] = None
        self.is_connected = False
        self.data_queue = queue.Queue()
        self.read_thread: Optional[threading.Thread] = None
        self.running = False

    @staticmethod
    def get_available_ports() -> List[str]:
        """사용 가능한 시리얼 포트 목록"""
        if not SERIAL_AVAILABLE:
            return []
        ports = serial.tools.list_ports.comports()
        return [port.device for port in ports]

    @staticmethod
    def get_available_ports_detail() -> List[Dict]:
        """사용 가능한 시리얼 포트 상세 목록"""
        if not SERIAL_AVAILABLE:
            return []
        ports = serial.tools.list_ports.comports()
        result = []
        for p in ports:
            result.append(
                {
                    "device": p.device,
                    "description": p.description or p.device,
                    "hwid": p.hwid or "",
                    "manufacturer": getattr(p, "manufacturer", "") or "",
                }
            )
        return sorted(result, key=lambda x: x["device"])

    @staticmethod
    def test_port(port: str, baudrate: int = 9600) -> Tuple[bool, str]:
        """포트 연결 테스트 (열기/닫기만 시도)"""
        if not SERIAL_AVAILABLE:
            return False, "pyserial 모듈이 설치되지 않았습니다"
        try:
            s = serial.Serial(
                port=port,
                baudrate=baudrate,
                bytesize=serial.EIGHTBITS,
                parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,
                timeout=1,
            )
            s.close()
            return True, f"{port} 연결 성공"
        except serial.SerialException as e:
            return False, f"{port} 연결 실패: {e}"
        except Exception as e:
            return False, f"{port} 오류: {e}"

    def connect(self) -> bool:
        """시리얼 포트 연결"""
        if not SERIAL_AVAILABLE:
            print("Serial module not available")
            return False

        try:
            self.serial = serial.Serial(
                port=self.port,
                baudrate=self.baudrate,
                bytesize=serial.EIGHTBITS,
                parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,
                timeout=1,
            )
            self.is_connected = True
            self.running = True
            self.read_thread = threading.Thread(target=self._read_loop, daemon=True)
            self.read_thread.start()
            print(f"Serial connected: {self.port} @ {self.baudrate}")
            return True
        except serial.SerialException as e:
            print(f"Serial connection error ({self.port}): {e}")
            return False
        except Exception as e:
            print(f"Serial connection error ({self.port}): {e}")
            return False

    def disconnect(self):
        """연결 해제"""
        self.running = False
        if self.serial and self.serial.is_open:
            self.serial.close()
        self.is_connected = False

    def _serial_log(self, msg: str):
        """시리얼 디버그 로그를 파일과 콘솔에 동시 출력 (flush 보장)"""
        import sys

        print(msg, flush=True)
        try:
            with open("serial_debug.log", "a", encoding="utf-8") as f:
                f.write(f"{msg}\n")
        except Exception:
            pass

    def _read_loop(self):
        """백그라운드 읽기 루프 — CR(\r) 종단 프로토콜 대응

        터치센서 데이터 형식 (실측 확인됨):
          1터치 = [0x00] + [숫자값\r]
          예: b'\x00' 이후 b'911\r'

        전략:
        - readline()으로 한 줄 단위 읽기
        - NULL/공백 등은 strip()으로 제거
        - 유효한 라인만 큐에 넣음
        """
        import time as _time

        self._serial_log("[Serial] _read_loop started (CR-terminated protocol)")

        while self.running and self.serial and self.serial.is_open:
            try:
                if self.serial.in_waiting > 0:
                    line = (
                        self.serial.readline().decode("ascii", errors="ignore").strip()
                    )
                    if line:
                        self._serial_log(f"[Serial] Received: {line!r}")
                        self.data_queue.put(line)
                else:
                    # 데이터 없으면 짧은 대기 (CPU 과부하 방지)
                    _time.sleep(0.01)
            except Exception as e:
                self._serial_log(f"[Serial ERROR] {e}")
                _time.sleep(0.05)

    def read_value(self, timeout: float = 1.0) -> Optional[float]:
        """측정값 읽기

        장비 데이터 형식 (실측 확인):
          '01A-000.0018'  →  축 prefix '01A' + 측정값 '-000.0018'
          - 앞 2~3자: 축 번호(숫자) + 채널(영문자)
          - 나머지: 부호 + 숫자값 (mm 단위)
        """
        import re

        try:
            data = self.data_queue.get(timeout=timeout)

            # 축/채널 prefix 제거 후 숫자값만 추출
            # 예: '01A-000.0018' → prefix='01A', value_str='-000.0018'
            #     '0A1-000.0019' → prefix='0A1', value_str='-000.0019'
            #     '01A+000.0018' → prefix='01A', value_str='+000.0018'
            # prefix 없는 순수 숫자 ('911', '-0.0018') 도 허용

            # 1) 부호(+/-)를 구분자로 prefix와 숫자값 분리
            match = re.match(r"^([^-+.]*?)([-+]\d+\.?\d*)$", data)
            if match and match.group(1):
                prefix = match.group(1)
                value_str = match.group(2)
                self._serial_log(f"[Serial] Prefix='{prefix}', Value='{value_str}'")
            else:
                # 2) 부호 없는 경우: 영문자 뒤의 숫자를 값으로 추출
                match2 = re.match(r"^([0-9A-Za-z]*[A-Za-z])(\d+\.?\d*)$", data)
                if match2:
                    prefix = match2.group(1)
                    value_str = match2.group(2)
                    self._serial_log(f"[Serial] Prefix='{prefix}', Value='{value_str}'")
                else:
                    # 3) prefix 없는 순수 숫자 데이터
                    value_str = data.replace(",", ".")

            raw_value = float(value_str)

            # 유효성 검증: 비정상적으로 큰 값 필터
            if abs(raw_value) > 99999:
                self._serial_log(f"[Serial] Out of range skipped: {raw_value}")
                return None

            self._serial_log(f"[Serial] Parsed value={raw_value:.6f}")
            return raw_value
        except queue.Empty:
            return None
        except ValueError:
            self._serial_log(f"[Serial] Parse error skipped: {data!r}")
            return None

    def send_command(self, cmd: str):
        """명령 전송"""
        if self.serial and self.serial.is_open:
            self.serial.write(cmd.encode("ascii"))

    def clear_buffer(self):
        """버퍼 초기화 - 큐와 시리얼 입력 버퍼 모두 비움"""
        # 큐 비우기
        while not self.data_queue.empty():
            try:
                self.data_queue.get_nowait()
            except queue.Empty:
                break
        # 시리얼 입력 버퍼 비우기
        if self.serial and self.serial.is_open:
            self.serial.reset_input_buffer()
        # 잠시 대기 후 다시 큐 비우기 (읽기 스레드가 방금 넣은 데이터 제거)
        import time

        time.sleep(0.1)
        while not self.data_queue.empty():
            try:
                self.data_queue.get_nowait()
            except queue.Empty:
                break


# =============================================================================
# 테스트 계산 로직
# =============================================================================


class TestCalculator:
    """측정값 분석 및 합격/불합격 판정"""

    @staticmethod
    def calculate_sigma(values: List[float]) -> float:
        """표준편차 계산 (샘플)"""
        if len(values) < 2:
            return 0.0
        return statistics.stdev(values)

    @staticmethod
    def calculate_range(values: List[float]) -> float:
        """범위 계산 (최대 - 최소)"""
        if not values:
            return 0.0
        return max(values) - min(values)

    @staticmethod
    def calculate_mean(values: List[float]) -> float:
        """평균 계산"""
        if not values:
            return 0.0
        return statistics.mean(values)

    @staticmethod
    def evaluate_axis_result(
        sigma: float,
        range_val: float,
        limits: LimitInfo,
        check_sigma: bool = True,
        check_range: bool = True,
    ) -> str:
        """축별 합격/불합격 판정"""
        if check_sigma and sigma > limits.mean_sigma:
            return "NG"
        if check_range and range_val > limits.worst_range:
            return "NG"
        return "OK"

    @staticmethod
    def evaluate_overall_result(
        mean_sigma: float,
        mean_range: float,
        worst_range: float,
        limits: LimitInfo,
        check_sigma: bool = True,
        check_range: bool = True,
    ) -> str:
        """전체 합격/불합격 판정"""
        if check_sigma and mean_sigma > limits.mean_sigma:
            return "NG"
        if check_range:
            if mean_range > limits.mean_range:
                return "NG"
            if worst_range > limits.worst_range:
                return "NG"
        return "OK"


# =============================================================================
# Excel 출력 모듈
# =============================================================================


class ExcelExporter:
    """Excel 파일 출력"""

    def __init__(self, template_path: str = None):
        self.template_path = template_path

    def export_result(
        self, result: TestResult, axis_results: List[AxisResult], output_path: str
    ) -> bool:
        """테스트 결과를 Excel로 출력"""
        if not EXCEL_AVAILABLE:
            print("Excel export not available")
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Results"

            # 헤더
            headers = [
                "Date",
                "Code",
                "Serial Number",
                "Operator",
                "Result",
                "Mean Sigma",
                "Mean Range",
                "Worst Sigma",
                "Worst Range",
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 데이터
            ws.cell(row=2, column=1, value=result.date.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=2, column=2, value=result.code)
            ws.cell(row=2, column=3, value=result.serial_number)
            ws.cell(row=2, column=4, value=result.operator)
            ws.cell(row=2, column=5, value=result.result)
            ws.cell(row=2, column=6, value=f"{result.mean_sigma:.3f}")
            ws.cell(row=2, column=7, value=f"{result.mean_range:.3f}")
            ws.cell(row=2, column=8, value=f"{result.worst_sigma:.3f}")
            ws.cell(row=2, column=9, value=f"{result.worst_range:.3f}")

            # 축별 결과
            ws2 = wb.create_sheet("Axis Results")
            axis_headers = ["Axis", "Direction", "Sigma", "Range", "Result", "NCycles"]
            for col, header in enumerate(axis_headers, 1):
                ws2.cell(row=1, column=col, value=header)

            for row_idx, axis_result in enumerate(axis_results, 2):
                ws2.cell(row=row_idx, column=1, value=axis_result.axis)
                ws2.cell(row=row_idx, column=2, value=axis_result.direction)
                ws2.cell(row=row_idx, column=3, value=f"{axis_result.sigma:.3f}")
                ws2.cell(row=row_idx, column=4, value=f"{axis_result.range_val:.3f}")
                ws2.cell(row=row_idx, column=5, value=axis_result.result)
                ws2.cell(row=row_idx, column=6, value=axis_result.ncycles)

            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Excel export error: {e}")
            return False


# =============================================================================
# PDF 인증서 출력 모듈
# =============================================================================


class CertificateGenerator:
    """PDF 인증서 생성 (form.xlsx 양식 기반)"""

    def __init__(self):
        # 리소스 경로
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.logo_path = os.path.join(self.script_dir, "resources", "logo.png")

    def generate(
        self,
        result: TestResult,
        axis_results: List[AxisResult],
        code_info: CodeInfo,
        output_path: str,
    ) -> bool:
        """인증서 PDF 생성 (form.xlsx 레이아웃과 동일)"""
        if not PDF_AVAILABLE:
            print("PDF generation not available - reportlab not installed")
            return False

        try:
            from reportlab.lib.pagesizes import A5
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
                Image,
            )
            from reportlab.lib.units import mm
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=10 * mm,
                leftMargin=10 * mm,
                topMargin=10 * mm,
                bottomMargin=10 * mm,
            )

            elements = []
            styles = getSampleStyleSheet()

            # 공통 스타일 (A5 축소)
            center_style = ParagraphStyle(
                "CenterCell",
                parent=styles["Normal"],
                fontSize=7,
                alignment=TA_CENTER,
            )
            bold_center_style = ParagraphStyle(
                "BoldCenterCell",
                parent=styles["Normal"],
                fontSize=7,
                alignment=TA_CENTER,
                fontName="Helvetica-Bold",
            )

            # ========== Row 1: 로고 (좌측) + INSPECTION SHEET (가운데) ==========
            logo_cell = ""
            if os.path.exists(self.logo_path):
                try:
                    logo_cell = Image(self.logo_path, width=25 * mm, height=25 * mm)
                except Exception as e:
                    print(f"Logo load error: {e}")
                    logo_cell = ""

            title_style = ParagraphStyle(
                "TitleCell",
                parent=styles["Heading1"],
                fontSize=16,
                alignment=TA_CENTER,
                leading=20,
            )
            title_para = Paragraph("<b>INSPECTION SHEET</b>", title_style)

            header_data = [[logo_cell, title_para]]
            header_table = Table(header_data, colWidths=[30 * mm, 150 * mm])
            header_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (0, 0), "LEFT"),
                        ("ALIGN", (1, 0), (1, 0), "CENTER"),
                    ]
                )
            )
            elements.append(header_table)
            elements.append(Spacer(1, 5 * mm))

            # ========== Probe Model / Code / Serial (셀 병합 및 가운데정렬) ==========
            probe_type = code_info.probe_type if code_info else ""
            info_data = [
                [
                    "Probe Model",
                    probe_type,
                    "Code",
                    result.code,
                    "Serial",
                    result.serial_number,
                ],
            ]
            info_table = Table(
                info_data,
                colWidths=[25 * mm, 45 * mm, 15 * mm, 40 * mm, 15 * mm, 40 * mm],
            )
            info_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
                        ("FONTNAME", (2, 0), (2, 0), "Helvetica-Bold"),
                        ("FONTNAME", (4, 0), (4, 0), "Helvetica-Bold"),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            elements.append(info_table)
            elements.append(Spacer(1, 5 * mm))

            # ========== TEST CYCLE DESCRIPTION ==========
            section_style = ParagraphStyle(
                "Section",
                parent=styles["Heading2"],
                fontSize=11,
                alignment=TA_LEFT,
                spaceAfter=3,
                spaceBefore=5,
                textColor=colors.black,
                backColor=colors.Color(0.9, 0.9, 0.9),
            )
            elements.append(Paragraph("TEST CYCLE DESCRIPTION", section_style))

            # ========== Cycle sequence 설명 (100 times 고정) ==========
            ncycles = 100
            if axis_results and len(axis_results) > 0:
                ncycles = axis_results[0].ncycles or 100

            cycle_desc_style = ParagraphStyle(
                "CycleDesc",
                parent=styles["Normal"],
                fontSize=7,
                alignment=TA_CENTER,
                spaceBefore=3,
                spaceAfter=3,
            )
            cycle_desc = f"Cycle sequence: X+ X- Y+ Y- Z- touch direction repeated {ncycles} times"
            elements.append(Paragraph(cycle_desc, styles["Normal"]))
            elements.append(Spacer(1, 3 * mm))

            # ========== Direction / Range 테이블 (셀 병합 및 가운데정렬) ==========
            dir_labels = ["Y-", "X+", "Y+", "X-"]

            # 축별 Range 데이터 (소수점 1자리)
            axis_ranges = []
            axis_2sigmas = []
            for i in range(4):
                if i < len(axis_results):
                    axis_ranges.append(f"{axis_results[i].range_val:.1f}")
                    axis_2sigmas.append(f"{2.0 * axis_results[i].sigma:.1f}")
                else:
                    axis_ranges.append("-")

            # Mean Range 계산
            mean_range_val = f"{result.mean_range:.1f}"

            # micron 단위를 데이터 셀에 병합 (9열 → 5열)
            def _fmt_micron(val: str) -> str:
                """데이터 값에 micron 단위 병합 (빈 값이면 그대로 반환)"""
                return f"{val} micron" if val and val != "-" else val

            axis_data = [
                [
                    "Direction",
                    dir_labels[0],
                    dir_labels[1],
                    dir_labels[2],
                    dir_labels[3],
                ],
                [
                    f"R({ncycles})={result.worst_range_limit:.1f}Micron",
                    _fmt_micron(axis_ranges[0]),
                    _fmt_micron(axis_ranges[1]),
                    _fmt_micron(axis_ranges[2]),
                    _fmt_micron(axis_ranges[3]),
                ],
                [
                    f"R({ncycles})={result.worst_range_limit:.1f}Micron",
                    f"{mean_range_val} micron",
                    "",
                    "",
                    "",
                ],
            ]

            col_w = 32 * mm
            axis_table = Table(
                axis_data,
                colWidths=[40 * mm, col_w, col_w, col_w, col_w],
            )
            axis_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.95, 0.95, 0.95)),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(axis_table)
            elements.append(Spacer(1, 3 * mm))

            # ========== Row 11-12: Un direct direction (Z-) — 5열 구조 ==========
            z_data = [
                ["Un direct direction", "Z-", "", "Dia", ""],
                ["", "Micron", "", "Micron", ""],
            ]
            z_table = Table(
                z_data,
                colWidths=[40 * mm, col_w, col_w, col_w, col_w],
            )
            z_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.95, 0.95, 0.95)),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(z_table)
            elements.append(Spacer(1, 8 * mm))

            # ========== 날짜 / TEST OK or NG / 작업자 (가운데정렬) ==========
            result_text = "TEST OK" if result.result == "OK" else "TEST NG"
            result_color = colors.darkgreen if result.result == "OK" else colors.red

            result_para_style = ParagraphStyle(
                "ResultText",
                parent=styles["Normal"],
                fontSize=12,
                alignment=TA_CENTER,
                textColor=result_color,
                fontName="Helvetica-Bold",
            )

            footer_data = [
                [
                    result.date.strftime("%Y-%m-%d"),
                    Paragraph(f"<b>{result_text}</b>", result_para_style),
                    f"operator: {result.operator}",
                ]
            ]
            footer_table = Table(footer_data, colWidths=[50 * mm, 60 * mm, 70 * mm])
            footer_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("ALIGN", (0, 0), (0, 0), "LEFT"),
                        ("ALIGN", (1, 0), (1, 0), "CENTER"),
                        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            elements.append(footer_table)

            # PDF 빌드
            doc.build(elements)
            print(f"PDF saved to: {output_path}")
            return True

        except Exception as e:
            import traceback

            print(f"PDF generation error: {e}")
            traceback.print_exc()
            return False


# =============================================================================
# 메인 GUI 애플리케이션
# =============================================================================


class MiniasApp:
    """MINIAS 메인 애플리케이션 (Tkinter GUI)"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("MiniAscott")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # 컴포넌트 초기화
        self.db = MiniasDatabase()
        self.serial = SerialCommunicator()
        self.calculator = TestCalculator()
        self.excel_exporter = ExcelExporter()
        self.cert_generator = CertificateGenerator()

        # 상태 변수
        self.is_testing = False
        self.is_paused = False
        self.current_axis = 1
        self.current_cycle = 1
        self.current_id = 0
        self.measurements: Dict[int, List[float]] = {}  # axis -> values
        self.axis_results: List[AxisResult] = []

        # 설정
        self.setup: Optional[SetupInfo] = None
        self.limits: Optional[LimitInfo] = None
        self.current_code_info: Optional[CodeInfo] = None

        # Tkinter 변수
        self.var_probe_type = tk.StringVar()
        self.var_serial_number = tk.StringVar()
        self.var_code = tk.StringVar()
        self.var_operator = tk.StringVar()
        self.var_check_range = tk.BooleanVar(value=True)
        self.var_check_sigma = tk.BooleanVar(value=True)
        self.var_naxis = tk.StringVar(value="4")
        self.var_ncycles = tk.StringVar(value="100")
        self.var_nworstdatums = tk.StringVar(value="1")
        self.var_id_col = tk.StringVar(value="")
        self.var_status = tk.StringVar(value="Ready")

        # INI 설정 로드
        self.config = self._load_config()

        # GUI 구성
        self._create_gui()

        # 데이터베이스 연결
        self.db.connect()

        # 초기 데이터 로드
        self._load_initial_data()

    def _load_config(self) -> Dict:
        """INI 설정 파일 로드"""
        config = {"port": "COM1", "baudrate": 9600, "working_dir": os.getcwd()}

        ini_path = os.path.join(os.path.dirname(__file__), "MINIAS.INI")
        if os.path.exists(ini_path):
            try:
                with open(ini_path, "r") as f:
                    for line in f:
                        if "=" in line:
                            key, value = line.strip().split("=", 1)
                            key = key.strip("[]")
                            if "Communication Port" in key and "Settings" not in key:
                                config["port"] = f"COM{value.strip()}"
                            elif "Settings Communication Port" in key:
                                parts = value.strip().split(",")
                                if parts:
                                    config["baudrate"] = int(parts[0])
                            elif "Working Directory" in key:
                                config["working_dir"] = value.strip()
            except Exception as e:
                print(f"Error loading config: {e}")

        # INI에 지정된 포트가 존재하지 않으면 사용 가능한 포트로 자동 전환
        available_ports = SerialCommunicator.get_available_ports()
        if available_ports and config["port"] not in available_ports:
            old_port = config["port"]
            config["port"] = available_ports[0]
            print(f"Warning: {old_port} not found. Auto-selected: {config['port']}")
            print(f"Available ports: {', '.join(available_ports)}")

        return config

    def _save_config(self):
        """INI 설정 파일 저장"""
        ini_path = os.path.join(os.path.dirname(__file__), "MINIAS.INI")
        # 포트 번호 추출 (COM3 -> 3, 알 수 없는 형식이면 그대로 저장)
        port_str = self.config.get("port", "COM1")
        if port_str.upper().startswith("COM"):
            port_num = port_str[3:]
        else:
            port_num = port_str

        baudrate = self.config.get("baudrate", 9600)
        working_dir = self.config.get("working_dir", os.getcwd())

        try:
            # 읽기 전용 파일인 경우 쓰기 가능하도록 변경
            import stat

            if os.path.exists(ini_path):
                file_stat = os.stat(ini_path)
                if not (file_stat.st_mode & stat.S_IWRITE):
                    os.chmod(ini_path, file_stat.st_mode | stat.S_IWRITE)

            # 기존 INI 파일의 추가 항목들을 보존
            extra_lines = []
            known_keys = {
                "Communication Port",
                "Settings Communication Port",
                "Working Directory",
            }
            if os.path.exists(ini_path):
                with open(ini_path, "r") as f:
                    for line in f:
                        if "=" in line:
                            key = line.strip().split("=", 1)[0].strip("[]")
                            if key not in known_keys:
                                extra_lines.append(line.rstrip())

            with open(ini_path, "w") as f:
                f.write(f"[Communication Port]={port_num}\n")
                f.write(f"[Settings Communication Port]={baudrate},N,8,1\n")
                f.write(f"[Working Directory]={working_dir}\n")
                for extra in extra_lines:
                    f.write(f"{extra}\n")
            print(f"Config saved: port={port_str}, baudrate={baudrate}")
        except Exception as e:
            print(f"Error saving config: {e}")

    def _create_gui(self):
        """GUI 구성요소 생성"""
        # 상단 버튼 프레임
        btn_frame = ttk.Frame(self.root, padding="5")
        btn_frame.pack(fill=tk.X)

        self.btn_start = ttk.Button(
            btn_frame, text="Start", command=self._on_start, width=12
        )
        self.btn_start.pack(side=tk.LEFT, padx=5)

        self.btn_stop = ttk.Button(
            btn_frame, text="STOP", command=self._on_stop, width=12
        )
        self.btn_stop.pack(side=tk.LEFT, padx=5)
        self.btn_print = ttk.Button(
            btn_frame, text="Save PDF", command=self._on_print_certificate
        )
        self.btn_print.pack(side=tk.LEFT, padx=5)

        self.btn_print_direct = ttk.Button(
            btn_frame, text="Print", command=self._on_print_direct
        )
        self.btn_print_direct.pack(side=tk.LEFT, padx=5)

        ttk.Label(btn_frame, textvariable=self.var_id_col, width=8).pack(
            side=tk.LEFT, padx=5
        )

        self.btn_exit = ttk.Button(
            btn_frame, text="EXIT", command=self._on_exit, width=12
        )
        self.btn_exit.pack(side=tk.RIGHT, padx=5)

        self.btn_settings = ttk.Button(
            btn_frame, text="Settings", command=self._on_settings, width=12
        )
        self.btn_settings.pack(side=tk.RIGHT, padx=5)

        # 입력 프레임
        input_frame = ttk.Frame(self.root, padding="5")
        input_frame.pack(fill=tk.X)

        # Probe Type
        ttk.Label(input_frame, text="Probe Type:").grid(
            row=0, column=0, sticky=tk.E, padx=5, pady=2
        )
        self.entry_probe_type = ttk.Entry(
            input_frame, textvariable=self.var_probe_type, width=25, state="readonly"
        )
        self.entry_probe_type.grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)

        # Code
        ttk.Label(input_frame, text="Code:").grid(
            row=0, column=2, sticky=tk.E, padx=5, pady=2
        )
        self.combo_code = ttk.Combobox(
            input_frame, textvariable=self.var_code, width=22
        )
        self.combo_code.grid(row=0, column=3, sticky=tk.W, padx=5, pady=2)
        self.combo_code.bind("<<ComboboxSelected>>", self._on_code_selected)

        # Serial Number
        ttk.Label(input_frame, text="Serial Number:").grid(
            row=1, column=0, sticky=tk.E, padx=5, pady=2
        )
        self.entry_serial = ttk.Entry(
            input_frame, textvariable=self.var_serial_number, width=25
        )
        self.entry_serial.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)

        # Checked by (Operator)
        ttk.Label(input_frame, text="Checked by:").grid(
            row=1, column=2, sticky=tk.E, padx=5, pady=2
        )
        self.combo_operator = ttk.Combobox(
            input_frame, textvariable=self.var_operator, width=22
        )
        self.combo_operator.grid(row=1, column=3, sticky=tk.W, padx=5, pady=2)

        # ID 조회 프레임
        id_frame = ttk.Frame(input_frame)
        id_frame.grid(row=2, column=0, columnspan=4, sticky=tk.W, pady=5)

        ttk.Label(id_frame, text="Serial/ID:").pack(side=tk.LEFT, padx=5)
        self.var_load_id = tk.StringVar()
        self.entry_load_id = ttk.Entry(
            id_frame, textvariable=self.var_load_id, width=10
        )
        self.entry_load_id.pack(side=tk.LEFT, padx=5)
        self.btn_load_id = ttk.Button(
            id_frame, text="Load", command=self._on_load_id, width=8
        )
        self.btn_load_id.pack(side=tk.LEFT, padx=5)

        # ID 목록 콤보박스
        ttk.Label(id_frame, text="or Select:").pack(side=tk.LEFT, padx=5)
        self.var_select_id = tk.StringVar()
        self.combo_id = ttk.Combobox(
            id_frame, textvariable=self.var_select_id, width=30, state="readonly"
        )
        self.combo_id.pack(side=tk.LEFT, padx=5)
        self.combo_id.bind("<<ComboboxSelected>>", self._on_id_selected)

        # 상태 표시 프레임 (시안색 배경)
        status_frame = tk.Frame(self.root, bg="#80FFFF", height=60)
        status_frame.pack(fill=tk.X, padx=10, pady=5)
        status_frame.pack_propagate(False)

        self.lbl_status = tk.Label(
            status_frame, textvariable=self.var_status, bg="#80FFFF", font=("Arial", 12)
        )
        self.lbl_status.pack(expand=True)

        # 체크 옵션 및 테스트 설정 프레임
        check_frame = ttk.Frame(self.root, padding="5")
        check_frame.pack(fill=tk.X)

        self.btn_limits = ttk.Button(
            check_frame, text="Limits", command=self._on_limits, width=10
        )
        self.btn_limits.pack(side=tk.LEFT, padx=5)

        ttk.Label(check_frame, text="Check").pack(side=tk.LEFT, padx=10)
        ttk.Checkbutton(check_frame, text="Range", variable=self.var_check_range).pack(
            side=tk.LEFT
        )
        ttk.Checkbutton(check_frame, text="Sigma", variable=self.var_check_sigma).pack(
            side=tk.LEFT
        )

        ttk.Label(check_frame, text="Test").pack(side=tk.LEFT, padx=20)

        ttk.Label(check_frame, text="NAxis:").pack(side=tk.LEFT, padx=5)
        ttk.Label(check_frame, textvariable=self.var_naxis, width=5).pack(side=tk.LEFT)

        ttk.Label(check_frame, text="NCycles:").pack(side=tk.LEFT, padx=5)
        ttk.Label(check_frame, textvariable=self.var_ncycles, width=5).pack(
            side=tk.LEFT
        )

        ttk.Label(check_frame, text="NWorstDatums:").pack(side=tk.LEFT, padx=5)
        ttk.Label(check_frame, textvariable=self.var_nworstdatums, width=5).pack(
            side=tk.LEFT
        )

        # Test Results 라벨
        ttk.Label(self.root, text="Test Results", font=("Arial", 10, "bold")).pack(
            anchor=tk.W, padx=10
        )

        # 결과 그리드 (Treeview)
        grid_frame = ttk.Frame(self.root, padding="5")
        grid_frame.pack(fill=tk.BOTH, expand=True, padx=5)

        columns = ("Axis", "Range", "Sigma", "Result", "2nd", "Result2", "Dir")
        self.tree = ttk.Treeview(grid_frame, columns=columns, show="headings", height=8)

        # 컬럼 설정
        col_widths = {
            "Axis": 60,
            "Range": 100,
            "Sigma": 100,
            "Result": 70,
            "2nd": 70,
            "Result2": 70,
            "Dir": 50,
        }
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths.get(col, 80), anchor=tk.CENTER)

        # 스크롤바
        scrollbar = ttk.Scrollbar(
            grid_frame, orient=tk.VERTICAL, command=self.tree.yview
        )
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 초기 행 추가 (Axis 1-4, Mean, Worst)
        self._init_grid_rows()

        # 하단 상태바
        self.statusbar = ttk.Label(
            self.root, text="Ready", relief=tk.SUNKEN, anchor=tk.W
        )
        self.statusbar.pack(side=tk.BOTTOM, fill=tk.X)

    def _init_grid_rows(self):
        """그리드 초기 행 생성"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Axis 1-4
        for i in range(1, 5):
            self.tree.insert("", "end", values=(str(i), "", "", "", "", "", str(i)))

        # Mean, Worst
        self.tree.insert("", "end", values=("Mean", "", "", "", "", "", ""))
        self.tree.insert("", "end", values=("Worst", "", "", "", "", "", ""))

    def _load_initial_data(self):
        """초기 데이터 로드"""
        # 코드 목록
        codes = self.db.get_code_list()
        self.combo_code["values"] = codes

        # 작업자 목록
        operators = self.db.get_operators()
        self.combo_operator["values"] = operators

        # 설정 로드
        self.setup = self.db.get_setup("ST")
        if self.setup:
            self.var_naxis.set(str(self.setup.naxis))
            self.var_ncycles.set(str(self.setup.ncycles))
            self.var_nworstdatums.set(str(self.setup.nworstdatums))
            self.var_check_range.set(self.setup.range_test)
            self.var_check_sigma.set(self.setup.sigma_test)

        # 한계값 로드
        self.limits = self.db.get_limits("ST")

        # 마지막 ID
        last_id = self.db.get_last_id()
        self.var_id_col.set(f"Id_col: {last_id}")

        # 최근 ID 목록
        self._refresh_id_list()

    def _on_code_selected(self, event=None):
        """코드 선택 시"""
        code = self.var_code.get()
        if code:
            self.current_code_info = self.db.get_code_info(code)
            if self.current_code_info:
                self.var_probe_type.set(self.current_code_info.probe_type)
                self.var_naxis.set(str(self.current_code_info.naxis))

    def _on_start(self):
        """테스트 시작"""
        if self.is_testing:
            return

        # 입력 검증
        if not self.var_code.get():
            messagebox.showwarning("Warning", "Please select a Code")
            return
        if not self.var_serial_number.get():
            messagebox.showwarning("Warning", "Please enter Serial Number")
            return
        if not self.var_operator.get():
            messagebox.showwarning("Warning", "Please select an Operator")
            return

        # 작업자 추가 (없으면)
        operator = self.var_operator.get()
        if operator not in self.db.get_operators():
            self.db.add_operator(operator)
            self.combo_operator["values"] = self.db.get_operators()

        # 측정 초기화
        self.measurements = {}
        self.axis_results = []
        self.current_axis = 1
        self.current_cycle = 1
        self.db.clear_measures()
        self._init_grid_rows()

        # 테스트 시작
        self.is_testing = True
        self.is_paused = False
        # Start 버튼 → PAUSE로 변경
        self.btn_start.config(text="PAUSE", command=self._on_pause)
        self.var_status.set(
            f"Testing... Axis {self.current_axis}, Cycle {self.current_cycle}"
        )

        # 시리얼 통신 연결 (설정된 경우)
        if SERIAL_AVAILABLE and self.config.get("port"):
            port = self.config["port"]
            baudrate = self.config["baudrate"]
            self.serial.port = port
            self.serial.baudrate = baudrate
            if self.serial.connect():
                self.statusbar.config(text=f"Connected to {port} @ {baudrate} bps")
            else:
                available = SerialCommunicator.get_available_ports()
                if available:
                    port_list = ", ".join(available)
                    self.statusbar.config(
                        text=f"{port} connection failed - Simulation mode. Available: {port_list}"
                    )
                    messagebox.showwarning(
                        "Serial Connection Failed",
                        f"'{port}' 포트에 연결할 수 없습니다.\n"
                        f"Simulation 모드로 실행됩니다.\n\n"
                        f"사용 가능한 포트: {port_list}\n\n"
                        f"Settings 메뉴에서 올바른 포트를 선택하세요.",
                    )
                else:
                    self.statusbar.config(
                        text="No serial ports detected - Simulation mode"
                    )
                    messagebox.showwarning(
                        "Serial Connection Failed",
                        "시리얼 포트가 감지되지 않습니다.\n"
                        "Simulation 모드로 실행됩니다.\n\n"
                        "장비가 연결되어 있는지 확인하세요.",
                    )
        elif not SERIAL_AVAILABLE:
            self.statusbar.config(text="pyserial not installed - Simulation mode")

        # 테스트 스레드 시작
        test_thread = threading.Thread(target=self._run_test, daemon=True)
        test_thread.start()

    def _run_test(self):
        """테스트 실행 (백그라운드)
        - 1터치 = 1사이클
        - 매 터치마다 측정값 디스플레이에 표기
        - 100사이클 완료 시 자동으로 다음 축으로 넘어가지 않음 (사용자 확인 대기)
        - NG 시 해당 축 재측정
        """
        import time

        naxis = int(self.var_naxis.get())
        ncycles = int(self.var_ncycles.get())

        axis = 1
        while axis <= naxis:
            self.current_axis = axis
            self.measurements[axis] = []

            # 시리얼 버퍼 초기화 (이전 축의 잔여 데이터 제거)
            if self.serial.is_connected:
                self.serial.clear_buffer()

            cycle = 1
            while cycle <= ncycles:
                if not self.is_testing:
                    return

                while self.is_paused:
                    time.sleep(0.1)
                    if not self.is_testing:
                        return

                self.current_cycle = cycle
                self.root.after(
                    0,
                    lambda a=axis, c=cycle: self.var_status.set(
                        f"Testing... Axis {a}, Cycle {c}/{ncycles}"
                    ),
                )

                # 측정값 읽기 - 1터치 = 1사이클
                # _read_loop가 NULL/빈줄 필터링 후 유효 데이터만 큐에 넣음
                # read_value는 큐에서 숫자 1개를 즉시 꺼냄 (딜레이 없음)
                value = None
                if self.serial.is_connected:
                    value = self.serial.read_value(timeout=2.0)

                if value is None:
                    # 시뮬레이션 모드
                    import random

                    value = random.gauss(0, 0.001)

                self.measurements[axis].append(value)
                self.db.save_measure(axis, cycle, value)

                # 매 터치마다 측정값을 상태바 + 그리드에 실시간 표시
                current_values = list(self.measurements[axis])
                cur_range = self.calculator.calculate_range(current_values)
                cur_sigma = (
                    self.calculator.calculate_sigma(current_values)
                    if len(current_values) >= 2
                    else 0.0
                )

                self.root.after(
                    0,
                    lambda a=axis, c=cycle, v=value, r=cur_range, s=cur_sigma: (
                        self.var_status.set(
                            f"Axis {a}, Cycle {c}/{ncycles} - Value: {v:.3f}"
                        ),
                        self._update_grid_row_live(a, c, ncycles, v, r, s),
                    ),
                )

                cycle += 1

            # 축 완료 - 결과 계산
            values = self.measurements[axis]
            sigma = self.calculator.calculate_sigma(values)
            range_val = self.calculator.calculate_range(values)

            # 결과 판정
            axis_result = "OK"
            if self.limits:
                axis_result = self.calculator.evaluate_axis_result(
                    sigma,
                    range_val,
                    self.limits,
                    self.var_check_sigma.get(),
                    self.var_check_range.get(),
                )

            # 축 결과 저장
            ar = AxisResult(
                axis=axis,
                sigma=sigma,
                range_val=range_val,
                result=axis_result,
                ncycles=ncycles,
                direction=str(axis),
            )
            self.axis_results.append(ar)

            # 그리드 업데이트
            self.root.after(
                0,
                lambda a=axis,
                r=range_val,
                s=sigma,
                res=axis_result: self._update_grid_row(a, r, s, res),
            )

        # 테스트 완료
        self.root.after(0, self._complete_test)

    def _update_grid_row_live(
        self,
        axis: int,
        cycle: int,
        ncycles: int,
        value: float,
        cur_range: float,
        cur_sigma: float,
    ):
        """매 터치마다 그리드 행 실시간 업데이트 (진행 중 상태)"""
        children = self.tree.get_children()
        if axis <= len(children):
            item = children[axis - 1]
            two_sigma = 2.0 * cur_sigma
            self.tree.item(
                item,
                values=(
                    str(axis),
                    f"{cur_range:.1f}",
                    f"{two_sigma:.1f}",
                    f"{cycle}/{ncycles}",
                    "",
                    "",
                    str(axis),
                ),
            )

    def _update_grid_row(self, axis: int, range_val: float, sigma: float, result: str):
        """그리드 행 업데이트 (2Sigma = 2 * sigma 표시) - 축 완료 시"""
        children = self.tree.get_children()
        if axis <= len(children):
            item = children[axis - 1]
            self.tree.item(
                item,
                values=(
                    str(axis),
                    f"{range_val:.1f}",
                    f"{sigma:.1f}",
                    result,
                    "",
                    "",
                    str(axis),
                ),
            )

    def _complete_test(self):
        """테스트 완료 처리"""
        self.is_testing = False
        self._reset_buttons()

        # 전체 결과 계산
        all_sigmas = [ar.sigma for ar in self.axis_results]
        all_ranges = [ar.range_val for ar in self.axis_results]

        mean_sigma = self.calculator.calculate_mean(all_sigmas) if all_sigmas else 0
        mean_range = self.calculator.calculate_mean(all_ranges) if all_ranges else 0
        worst_sigma = max(all_sigmas) if all_sigmas else 0
        worst_range = max(all_ranges) if all_ranges else 0

        # 전체 판정
        overall_result = "OK"
        if self.limits:
            overall_result = self.calculator.evaluate_overall_result(
                mean_sigma,
                mean_range,
                worst_range,
                self.limits,
                self.var_check_sigma.get(),
                self.var_check_range.get(),
            )

        # Mean, Worst 행 업데이트 (2Sigma = 2 * sigma 표시)
        children = self.tree.get_children()
        if len(children) >= 6:
            # Mean 행
            self.tree.item(
                children[4],
                values=(
                    "Mean",
                    f"{mean_range:.1f}",
                    f"{mean_sigma:.1f}",
                    "",
                    "",
                    "",
                    "",
                ),
            )
            # Worst 행
            self.tree.item(
                children[5],
                values=(
                    "Worst",
                    f"{worst_range:.1f}",
                    f"{worst_sigma:.1f}",
                    overall_result,
                    "",
                    "",
                    "",
                ),
            )

        # 결과 저장
        result = TestResult(
            date=datetime.now(),
            code=self.var_code.get(),
            serial_number=self.var_serial_number.get(),
            operator=self.var_operator.get(),
            test_type="ST",
            result=overall_result,
            mean_sigma=mean_sigma,
            mean_range=mean_range,
            worst_sigma=worst_sigma,
            worst_range=worst_range,
            mean_sigma_limit=self.limits.mean_sigma if self.limits else 0,
            mean_range_limit=self.limits.mean_range if self.limits else 0,
            worst_range_limit=self.limits.worst_range if self.limits else 0,
        )

        self.current_id = self.db.save_test_result(result)
        self.var_id_col.set(f"Id_col: {self.current_id}")

        # 축별 결과 저장
        for ar in self.axis_results:
            ar.id_col = self.current_id
            self.db.save_axis_result(ar)

        # 샘플 데이터 저장
        for axis, values in self.measurements.items():
            for cycle, value in enumerate(values, 1):
                self.db.save_sample(self.current_id, axis, cycle, value)

        # 상태 업데이트
        result_text = "PASS" if overall_result == "OK" else "FAIL"
        self.var_status.set(f"Test Complete - {result_text}")
        self.statusbar.config(text=f"Test saved with ID: {self.current_id}")

        # ID 목록 새로고침
        self._refresh_id_list()

        # 시리얼 연결 해제
        self.serial.disconnect()

    def _reset_buttons(self):
        """버튼 상태 초기화 — Start 복원"""
        self.btn_start.config(text="Start", command=self._on_start, state="normal")

    def _on_pause(self):
        """일시 정지/재개 — Start 버튼이 PAUSE/RESUME으로 토글"""
        if not self.is_testing:
            return

        self.is_paused = not self.is_paused
        if self.is_paused:
            self.btn_start.config(text="RESUME")
            self.var_status.set("PAUSED")
        else:
            self.btn_start.config(text="PAUSE")
            self.var_status.set(f"Testing... Axis {self.current_axis}")

    def _on_stop(self):
        """테스트 중단 — 현재까지의 결과를 저장하고 테스트 종료"""
        if not self.is_testing:
            return

        confirm = messagebox.askyesno(
            "Stop Test",
            "Stop the current test?\nResults measured so far will be saved.",
        )
        if not confirm:
            return

        # 테스트 중단
        self.is_testing = False
        self.is_paused = False

        # 현재까지 완료된 축의 결과 계산
        # (measurements에 데이터가 있는 축만 결과 생성)
        for axis, values in self.measurements.items():
            if not values:
                continue

            # 이미 axis_results에 있는 축은 건너뜀
            if any(ar.axis == axis for ar in self.axis_results):
                continue

            sigma = self.calculator.calculate_sigma(values)
            range_val = self.calculator.calculate_range(values)

            axis_result_str = "OK"
            if self.limits:
                axis_result_str = self.calculator.evaluate_axis_result(
                    sigma,
                    range_val,
                    self.limits,
                    self.var_check_sigma.get(),
                    self.var_check_range.get(),
                )

            ar = AxisResult(
                axis=axis,
                sigma=sigma,
                range_val=range_val,
                result=axis_result_str,
                ncycles=len(values),
                direction=str(axis),
            )
            self.axis_results.append(ar)

        # 결과가 있으면 저장
        if self.axis_results:
            self._complete_test()
            self.var_status.set("Test STOPPED - Results saved")
        else:
            self.var_status.set("Test STOPPED - No results to save")

        self._reset_buttons()
        self.serial.disconnect()

    def _stop_and_save_current(self):
        """Pause 상태에서 SavePDF 시 호출 — 현재까지 데이터를 저장"""
        # 테스트 중단
        self.is_testing = False
        self.is_paused = False

        # 현재까지 완료된 축 결과 계산
        for axis, values in self.measurements.items():
            if not values:
                continue
            if any(ar.axis == axis for ar in self.axis_results):
                continue

            sigma = self.calculator.calculate_sigma(values)
            range_val = self.calculator.calculate_range(values)

            axis_result_str = "OK"
            if self.limits:
                axis_result_str = self.calculator.evaluate_axis_result(
                    sigma,
                    range_val,
                    self.limits,
                    self.var_check_sigma.get(),
                    self.var_check_range.get(),
                )

            ar = AxisResult(
                axis=axis,
                sigma=sigma,
                range_val=range_val,
                result=axis_result_str,
                ncycles=len(values),
                direction=str(axis),
            )
            self.axis_results.append(ar)

        if self.axis_results:
            self._complete_test()

        self._reset_buttons()
        self.serial.disconnect()

    def _on_print_certificate(self):
        """인증서 출력 — 테스트 완료 또는 Pause/Stop 상태에서 사용 가능"""
        # Pause 상태에서 호출된 경우: 현재까지 결과를 먼저 저장
        if self.is_testing and self.is_paused and self.current_id == 0:
            save_now = messagebox.askyesno(
                "Save Current Results",
                "Test is paused. Save current results and generate PDF?",
            )
            if not save_now:
                return
            # 현재까지 데이터로 결과 저장
            self._stop_and_save_current()

        if self.current_id == 0:
            messagebox.showwarning("Warning", "No test result to print")
            return

        # Serial Number를 기본 파일명으로 사용
        serial_number = (
            self.var_serial_number.get().strip() or f"Certificate_{self.current_id}"
        )

        # 파일 저장 대화상자
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=f"Certificate_{self.current_id}.pdf",
        )

        if file_path:
            result = self.db.get_test_result(self.current_id)
            axis_results = self.db.get_axis_results(self.current_id)

        if not result:
            messagebox.showerror("Error", "Test result not found in database")
            return

        if not axis_results:
            messagebox.showerror(
                "Error",
                f"No axis results found for test ID {self.current_id}.\n"
                "The test may not have completed properly.",
            )
            return

        # 코드 정보 가져오기 (current_code_info가 없으면 DB에서 조회)
        code_info = self.current_code_info
        if code_info is None and result.code:
            code_info = self.db.get_code_info(result.code)
        if code_info is None:
            code_info = CodeInfo(code=result.code, probe_type="")

            try:
                success = self.cert_generator.generate(
                    result, axis_results, code_info, file_path
                )
                if success:
                    messagebox.showinfo("Success", f"Certificate saved to {file_path}")
                else:
                    messagebox.showerror("Error", "Failed to generate certificate")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate certificate: {e}")

    def _on_delete_result(self):
        """현재 로드된 테스트 결과(성적서) 삭제"""
        if self.current_id == 0:
            messagebox.showwarning("Warning", "No test result loaded to delete")
            return

        # 확인 대화상자
        serial = self.var_serial_number.get().strip() or "N/A"
        confirm = messagebox.askyesno(
            "Confirm Delete",
            f"Delete test result?\n\n"
            f"  ID: {self.current_id}\n"
            f"  Serial: {serial}\n\n"
            f"This action cannot be undone.",
        )
        if not confirm:
            return

        try:
            self.db.delete_test_result(self.current_id)
            messagebox.showinfo("Deleted", f"Test result ID {self.current_id} deleted.")

            # UI 초기화
            self.current_id = 0
            self.var_serial_number.set("")
            self.var_load_id.set("")
            self.var_status.set("Test result deleted.")

            # 결과 테이블 클리어
            for item in self.tree_results.get_children():
                self.tree_results.delete(item)

            # ID 목록 새로고침
            self._refresh_id_list()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete test result: {e}")

    def _refresh_id_list(self):
        """ID 목록 새로고침"""
        recent_ids = self.db.get_recent_ids(50)
        id_display = [
            f"{id_col} - {code} ({serial})" for id_col, code, serial in recent_ids
        ]
        self.combo_id["values"] = id_display

    def _on_load_id(self):
        """Serial Number 우선으로 테스트 결과 로드 (fallback: ID 검색)"""
        search_text = self.var_load_id.get().strip()
        if not search_text:
            messagebox.showwarning("Warning", "Please enter a Serial Number or ID")
            return

        # 1) Serial Number로 먼저 검색
        results = self.db.search_by_serial_number(search_text)

        # 2) Serial Number 결과가 없고 숫자인 경우 → ID로 fallback 검색
        if not results:
            try:
                id_col = int(search_text)
                self._load_test_result(id_col)
                return
            except ValueError:
                pass
            messagebox.showwarning(
                "Warning", f"No test result found for: {search_text}"
            )
            return

        if len(results) == 1:
            # 결과가 하나면 바로 로드
            self._load_test_result(results[0][0])
        else:
            # 여러 결과가 있으면 선택 대화상자
            dialog = tk.Toplevel(self.root)
            dialog.title("Select Test Result")
            dialog.geometry("400x300")
            dialog.transient(self.root)
            dialog.grab_set()

            ttk.Label(
                dialog,
                text=f"Found {len(results)} results for '{search_text}':",
                font=("Arial", 10),
            ).pack(pady=10, padx=10)

            listbox = tk.Listbox(dialog, width=50, height=10)
            listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

            for id_col, code, serial in results:
                listbox.insert(
                    tk.END, f"ID: {id_col} - Code: {code} - Serial: {serial}"
                )

            def on_select():
                sel = listbox.curselection()
                if sel:
                    selected_id = results[sel[0]][0]
                    dialog.destroy()
                    self._load_test_result(selected_id)

            ttk.Button(dialog, text="Load", command=on_select, width=10).pack(pady=10)
            dialog.wait_window()

    def _on_id_selected(self, event=None):
        """콤보박스에서 ID 선택"""
        selected = self.var_select_id.get()
        if selected:
            try:
                id_col = int(selected.split(" - ")[0])
                self._load_test_result(id_col)
            except (ValueError, IndexError):
                pass

    def _load_test_result(self, id_col: int):
        """테스트 결과 로드 및 화면에 표시"""
        result = self.db.get_test_result(id_col)
        if not result:
            messagebox.showwarning("Warning", f"No test result found with ID: {id_col}")
            return

        axis_results = self.db.get_axis_results(id_col)

        # 화면에 데이터 표시
        self.var_code.set(result.code)
        self.var_serial_number.set(result.serial_number)
        self.var_operator.set(result.operator)

        # 코드 정보 로드
        if result.code:
            self.current_code_info = self.db.get_code_info(result.code)
            if self.current_code_info:
                self.var_probe_type.set(self.current_code_info.probe_type)

        # 그리드 업데이트
        self._init_grid_rows()
        children = self.tree.get_children()

        for ar in axis_results:
            axis_num = int(ar.axis) if ar.axis else 0
            if axis_num > 0 and axis_num <= len(children):
                item = children[axis_num - 1]
                self.tree.item(
                    item,
                    values=(
                        str(axis_num),
                        f"{ar.range_val:.1f}",
                        f"{ar.sigma:.1f}",
                        ar.result,
                        "",
                        "",
                        ar.direction or str(axis_num),
                    ),
                )

        # Mean, Worst 업데이트 (2Sigma 표시)
        if len(children) >= 6:
            self.tree.item(
                children[4],
                values=(
                    "Mean",
                    f"{result.mean_range:.1f}",
                    f"{result.mean_sigma:.1f}",
                    "",
                    "",
                    "",
                    "",
                ),
            )
            self.tree.item(
                children[5],
                values=(
                    "Worst",
                    f"{result.worst_range:.1f}",
                    f"{result.worst_sigma:.1f}",
                    result.result,
                    "",
                    "",
                    "",
                ),
            )

        # 상태 업데이트
        self.current_id = id_col
        self.axis_results = axis_results
        self.var_id_col.set(f"Id_col: {id_col}")
        result_text = "PASS" if result.result == "OK" else "FAIL"
        self.var_status.set(f"Loaded ID: {id_col} - {result_text}")
        self.statusbar.config(
            text=f"Test result loaded: {result.date.strftime('%Y-%m-%d %H:%M:%S')}"
        )

    def _on_print_direct(self):
        """PDF를 생성하고 바로 프린터로 출력"""
        if self.current_id == 0:
            messagebox.showwarning("Warning", "No test result to print")
            return

        result = self.db.get_test_result(self.current_id)
        axis_results = self.db.get_axis_results(self.current_id)

        if result:
            # 코드 정보 가져오기
            code_info = self.current_code_info
            if code_info is None and result.code:
                code_info = self.db.get_code_info(result.code)
            if code_info is None:
                code_info = CodeInfo(code=result.code, probe_type="")

            # 임시 파일로 PDF 생성
            import tempfile

            temp_dir = tempfile.gettempdir()
            temp_path = os.path.join(temp_dir, f"Certificate_{self.current_id}.pdf")

            try:
                success = self.cert_generator.generate(
                    result, axis_results, code_info, temp_path
                )
                if success:
                    # PDF 파일 열기 (사용자가 직접 인쇄)
                    try:
                        os.startfile(temp_path)
                        self.statusbar.config(
                            text=f"PDF opened: Certificate_{self.current_id}.pdf - Please print from the viewer"
                        )
                        messagebox.showinfo(
                            "Print",
                            "PDF 파일이 열렸습니다.\nPDF 뷰어에서 Ctrl+P를 눌러 인쇄하세요.",
                        )
                    except OSError as e:
                        # PDF 뷰어가 없는 경우 파일 위치 알림
                        messagebox.showinfo(
                            "Print",
                            f"PDF 파일이 생성되었습니다:\n{temp_path}\n\n파일을 열어서 인쇄해주세요.",
                        )
                        self.statusbar.config(text=f"PDF saved: {temp_path}")
                else:
                    messagebox.showerror("Error", "Failed to generate certificate")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to print certificate: {e}")

    def _on_settings(self):
        """통신 설정 대화상자"""
        SettingsDialog(self.root, self.config, self._apply_settings)

    def _apply_settings(self, new_config: Dict):
        """설정 적용 및 INI 파일 저장"""
        self.config.update(new_config)
        self.serial.port = self.config.get("port", "COM1")
        self.serial.baudrate = self.config.get("baudrate", 9600)
        self._save_config()
        self.statusbar.config(
            text=f"Settings updated: {self.config['port']} @ {self.config['baudrate']} bps"
        )

    def _on_limits(self):
        """한계값 설정 대화상자"""
        LimitsDialog(self.root, self.db, self.limits)
        # 한계값 다시 로드
        self.limits = self.db.get_limits("ST")

    def _on_exit(self):
        """프로그램 종료"""
        if self.is_testing:
            if not messagebox.askyesno("Confirm", "Test in progress. Exit anyway?"):
                return
            self.is_testing = False

        self.serial.disconnect()
        self.db.close()
        self.root.destroy()

    def run(self):
        """애플리케이션 실행"""
        self.root.protocol("WM_DELETE_WINDOW", self._on_exit)
        self.root.mainloop()


# =============================================================================
# 한계값 설정 대화상자
# =============================================================================


class LimitsDialog:
    """한계값 설정 대화상자"""

    def __init__(self, parent, db: MiniasDatabase, limits: Optional[LimitInfo]):
        self.db = db
        self.limits = limits or LimitInfo()

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Limits Settings")
        self.dialog.geometry("400x280")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # 변수
        self.var_mean_sigma = tk.StringVar(value=f"{self.limits.mean_sigma:.3f}")
        self.var_mean_range = tk.StringVar(value=f"{self.limits.mean_range:.3f}")
        self.var_worst_range = tk.StringVar(value=f"{self.limits.worst_range:.3f}")

        # GUI
        frame = ttk.Frame(self.dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Test Type: ST", font=("Arial", 10, "bold")).grid(
            row=0, column=0, columnspan=2, pady=10
        )

        ttk.Label(frame, text="Mean Sigma Limit:").grid(
            row=1, column=0, sticky=tk.E, pady=5
        )
        ttk.Entry(frame, textvariable=self.var_mean_sigma, width=15).grid(
            row=1, column=1, pady=5
        )

        ttk.Label(frame, text="Mean Range Limit:").grid(
            row=2, column=0, sticky=tk.E, pady=5
        )
        ttk.Entry(frame, textvariable=self.var_mean_range, width=15).grid(
            row=2, column=1, pady=5
        )

        ttk.Label(frame, text="Worst Range Limit:").grid(
            row=3, column=0, sticky=tk.E, pady=5
        )
        ttk.Entry(frame, textvariable=self.var_worst_range, width=15).grid(
            row=3, column=1, pady=5
        )

        # 버튼
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="Save", command=self._save, width=10).pack(
            side=tk.LEFT, padx=10
        )
        ttk.Button(
            btn_frame, text="Cancel", command=self.dialog.destroy, width=10
        ).pack(side=tk.LEFT, padx=10)

        self.dialog.wait_window()

    def _save(self):
        """저장"""
        try:
            self.limits.test_type = "ST"
            self.limits.mean_sigma = float(self.var_mean_sigma.get())
            self.limits.mean_range = float(self.var_mean_range.get())
            self.limits.worst_range = float(self.var_worst_range.get())

            self.db.save_limits(self.limits)
            messagebox.showinfo("Success", "Limits saved successfully")
            self.dialog.destroy()
        except ValueError:
            messagebox.showerror("Error", "Invalid number format")


# =============================================================================
# 통신 설정 대화상자
# =============================================================================


class SettingsDialog:
    """통신 설정 대화상자"""

    BAUDRATES = ["9600", "19200", "38400", "57600", "115200"]

    def __init__(self, parent, config: Dict, callback):
        self.config = config.copy()
        self.callback = callback
        self.port_details: List[Dict] = []

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Communication Settings")
        self.dialog.geometry("520x520")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # 변수
        self.var_port = tk.StringVar(value=self.config.get("port", "COM1"))
        self.var_baudrate = tk.StringVar(value=str(self.config.get("baudrate", 9600)))

        # GUI
        frame = ttk.Frame(self.dialog, padding="15")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            frame, text="Communication Settings", font=("Arial", 11, "bold")
        ).grid(row=0, column=0, columnspan=2, pady=(0, 10))

        # COM Port
        ttk.Label(frame, text="COM Port:").grid(
            row=1, column=0, sticky=tk.E, pady=5, padx=10
        )

        # 사용 가능한 포트 목록 가져오기
        available_ports = SerialCommunicator.get_available_ports()
        if not available_ports:
            available_ports = [
                "COM1",
                "COM2",
                "COM3",
                "COM4",
                "COM5",
                "COM6",
                "COM7",
                "COM8",
            ]

        self.combo_port = ttk.Combobox(
            frame, textvariable=self.var_port, width=15, values=available_ports
        )
        self.combo_port.grid(row=1, column=1, sticky=tk.W, pady=5)
        self.combo_port.bind("<<ComboboxSelected>>", self._on_port_selected)

        # Baudrate
        ttk.Label(frame, text="Baudrate:").grid(
            row=2, column=0, sticky=tk.E, pady=5, padx=10
        )
        self.combo_baudrate = ttk.Combobox(
            frame,
            textvariable=self.var_baudrate,
            width=15,
            values=self.BAUDRATES,
            state="readonly",
        )
        self.combo_baudrate.grid(row=2, column=1, sticky=tk.W, pady=5)

        # 포트 상세정보 라벨
        self.lbl_port_desc = ttk.Label(frame, text="", foreground="gray")
        self.lbl_port_desc.grid(row=3, column=0, columnspan=2, pady=5)

        # --- 포트 목록 Treeview ---
        tree_label = ttk.Label(frame, text="Detected Ports:", font=("Arial", 9, "bold"))
        tree_label.grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=(10, 2))

        tree_frame = ttk.Frame(frame)
        tree_frame.grid(row=5, column=0, columnspan=2, sticky="nsew", pady=2)

        self.port_tree = ttk.Treeview(
            tree_frame,
            columns=("port", "description"),
            show="headings",
            height=5,
            selectmode="browse",
        )
        self.port_tree.heading("port", text="Port")
        self.port_tree.heading("description", text="Description")
        self.port_tree.column("port", width=80, anchor=tk.W)
        self.port_tree.column("description", width=380, anchor=tk.W)
        self.port_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.port_tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        scrollbar = ttk.Scrollbar(
            tree_frame, orient=tk.VERTICAL, command=self.port_tree.yview
        )
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.port_tree.configure(yscrollcommand=scrollbar.set)

        # 포트 상태 표시
        self.lbl_status = ttk.Label(frame, text="", foreground="gray")
        self.lbl_status.grid(row=6, column=0, columnspan=2, pady=5)

        # --- 버튼 행 1: Refresh + Test ---
        action_frame = ttk.Frame(frame)
        action_frame.grid(row=7, column=0, columnspan=2, pady=5)

        ttk.Button(
            action_frame, text="Refresh Ports", command=self._refresh_ports, width=14
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            action_frame,
            text="Test Connection",
            command=self._test_connection,
            width=14,
        ).pack(side=tk.LEFT, padx=5)

        # --- 버튼 행 2: Apply + Cancel ---
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=8, column=0, columnspan=2, pady=(10, 0))

        ttk.Button(btn_frame, text="Apply", command=self._apply, width=10).pack(
            side=tk.LEFT, padx=10
        )
        ttk.Button(
            btn_frame, text="Cancel", command=self.dialog.destroy, width=10
        ).pack(side=tk.LEFT, padx=10)

        # 초기 데이터 로드
        self._update_port_tree()
        self._on_port_selected()

        self.dialog.wait_window()

    def _get_port_list(self) -> List[str]:
        """사용 가능한 포트 목록 (device 이름만)"""
        self.port_details = SerialCommunicator.get_available_ports_detail()
        if self.port_details:
            return [p["device"] for p in self.port_details]
        return ["COM1", "COM2", "COM3", "COM4"]

    def _update_port_tree(self):
        """포트 상세 Treeview 업데이트"""
        for item in self.port_tree.get_children():
            self.port_tree.delete(item)

        self.port_details = SerialCommunicator.get_available_ports_detail()
        if self.port_details:
            for p in self.port_details:
                self.port_tree.insert(
                    "", tk.END, values=(p["device"], p["description"])
                )
            self.lbl_status.config(
                text=f"{len(self.port_details)} port(s) detected", foreground="green"
            )
        else:
            self.lbl_status.config(
                text="No serial ports detected. Check device connection.",
                foreground="red",
            )

    def _on_port_selected(self, event=None):
        """포트 선택 시 상세정보 표시"""
        selected_port = self.var_port.get()
        for p in self.port_details:
            if p["device"] == selected_port:
                desc = p["description"]
                if p["manufacturer"]:
                    desc += f" [{p['manufacturer']}]"
                self.lbl_port_desc.config(text=desc)
                return
        self.lbl_port_desc.config(text="")

    def _on_tree_select(self, event=None):
        """Treeview에서 포트 선택 시 combobox에 반영"""
        selection = self.port_tree.selection()
        if selection:
            item = self.port_tree.item(selection[0])
            port_device = item["values"][0]
            self.var_port.set(port_device)
            self._on_port_selected()

    def _refresh_ports(self):
        """포트 목록 새로고침"""
        available_ports = SerialCommunicator.get_available_ports()
        if not available_ports:
            available_ports = [
                "COM1",
                "COM2",
                "COM3",
                "COM4",
                "COM5",
                "COM6",
                "COM7",
                "COM8",
            ]
        self.combo_port["values"] = available_ports
        self._update_port_tree()

    def _test_connection(self):
        """선택된 포트의 연결 테스트"""
        port = self.var_port.get()
        baudrate = int(self.var_baudrate.get())
        if not port:
            messagebox.showwarning("Warning", "포트를 선택하세요.")
            return

        self.lbl_status.config(text=f"Testing {port}...", foreground="orange")
        self.dialog.update()

        success, message = SerialCommunicator.test_port(port, baudrate)
        if success:
            self.lbl_status.config(text=f"{port}: {message}", foreground="green")
        else:
            self.lbl_status.config(text=f"{port}: {message}", foreground="red")

    def _apply(self):
        """설정 적용"""
        self.config["port"] = self.var_port.get()
        self.config["baudrate"] = int(self.var_baudrate.get())

        if self.callback:
            self.callback(self.config)

        messagebox.showinfo(
            "Success",
            f"Settings applied:\nPort: {self.config['port']}\nBaudrate: {self.config['baudrate']}",
        )
        self.dialog.destroy()


# =============================================================================
# 진입점
# =============================================================================


def main():
    """메인 함수"""
    app = MiniasApp()
    app.run()


if __name__ == "__main__":
    main()
