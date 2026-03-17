"""Excel 출력 모듈"""

from typing import List

# Excel 출력
try:
    from openpyxl import Workbook, load_workbook

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export disabled.")

from minias.models import TestResult, AxisResult


class ExcelExporter:
    """Excel 파일 출력"""

    def __init__(self, template_path: str = None):
        self.template_path = template_path

    def export_result(
        self, result: TestResult, axis_results: List[AxisResult], output_path: str
    ) -> bool:
        """테스트 결과를 Excel로 출력"""
        if not EXCEL_AVAILABLE:
            print("Excel export not available")
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Results"

            # 헤더
            headers = [
                "Date",
                "Code",
                "Serial Number",
                "Operator",
                "Result",
                "Mean 2SIGMA",
                "Mean Range",
                "Worst 2SIGMA",
                "Worst Range",
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 데이터
            ws.cell(row=2, column=1, value=result.date.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=2, column=2, value=result.code)
            ws.cell(row=2, column=3, value=result.serial_number)
            ws.cell(row=2, column=4, value=result.operator)
            ws.cell(row=2, column=5, value=result.result)
            ws.cell(row=2, column=6, value=f"{result.mean_sigma * 2000:.1f}")
            ws.cell(row=2, column=7, value=f"{result.mean_range * 1000:.1f}")
            ws.cell(row=2, column=8, value=f"{result.worst_sigma * 2000:.1f}")
            ws.cell(row=2, column=9, value=f"{result.worst_range * 1000:.1f}")

            # 축별 결과
            ws2 = wb.create_sheet("Axis Results")
            axis_headers = ["Axis", "Direction", "2SIGMA", "Range", "Result", "NCycles"]
            for col, header in enumerate(axis_headers, 1):
                ws2.cell(row=1, column=col, value=header)

            for row_idx, axis_result in enumerate(axis_results, 2):
                ws2.cell(row=row_idx, column=1, value=axis_result.axis)
                ws2.cell(row=row_idx, column=2, value=axis_result.direction)
                ws2.cell(row=row_idx, column=3, value=f"{axis_result.sigma * 2000:.1f}")
                ws2.cell(
                    row=row_idx, column=4, value=f"{axis_result.range_val * 1000:.1f}"
                )
                ws2.cell(row=row_idx, column=5, value=axis_result.result)
                ws2.cell(row=row_idx, column=6, value=axis_result.ncycles)

            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Excel export error: {e}")
            return False
